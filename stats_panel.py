"""
stats_panel.py — Панель «Статистика групп» MAX POST.

Загружает данные о группах напрямую через GREEN-API (независимо от внешнего сервера):
  - getChatHistory (count=1) → время последней активности (1 запрос на группу, кэш 6 ч)
  - getGroupData             → название + кол-во участников (~190 запросов батчами, кэш 1 ч)
"""
from __future__ import annotations

import json
import logging
import os
import re
import sys
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from PyQt6.QtCore import QThread, QTimer, QUrl, Qt, pyqtSignal
from PyQt6.QtGui import QColor, QDesktopServices, QFont
from PyQt6.QtWidgets import (
    QApplication, QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QLineEdit, QMenu, QProgressBar, QPushButton, QStackedWidget, QTableWidget,
    QTableWidgetItem, QVBoxLayout, QWidget,
)

try:
    from PyQt6.QtWebEngineWidgets import QWebEngineView
    from PyQt6.QtWebEngineCore import QWebEnginePage
    _WEB_ENGINE_AVAILABLE = True
except Exception as _web_err:
    import logging as _logging
    _logging.getLogger(__name__).warning("PyQt6-WebEngine недоступен: %s", _web_err)
    _WEB_ENGINE_AVAILABLE = False

_WEB_REPORT_URL = "https://bot-dev.gkh.spb.ru/gks2vyb-report.php"

from env_utils import get_env_path

_log = logging.getLogger(__name__)

_AUTO_REFRESH_MS        = 5 * 60 * 1000   # авто-обновление каждые 5 минут
_REQUEST_DELAY          = 1.1             # секунд между запросами (лимит GREEN-API: 1 req/s)
_GROUP_CACHE_TTL        = 3600            # секунд — кэш названия/участников группы (1 час)
_ACTIVITY_CACHE_TTL     = 6 * 3600       # секунд — кэш последней активности группы (6 часов)

# Индексы колонок таблицы
_COL_NAME    = 0
_COL_MEMBERS = 1
_COL_TIME    = 2
_COL_DELTA   = 3
_COL_LINK    = 4


def _extract_chat_id(raw: str) -> str:
    """Извлекает числовой chat_id из URL или возвращает строку как есть.

    Примеры:
      https://web.max.ru/-69098384919255  → -69098384919255
      https://web.max.ru/-69098384919255/ → -69098384919255
      -69098384919255                     → -69098384919255  (без изменений)
    """
    if raw.startswith("http"):
        m = re.search(r"(-\d+)/?$", raw)
        if m:
            return m.group(1)
    return raw


def _resolve_excel_path() -> Path:
    base = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
    return base / "max_address.xlsx"


def _cache_path() -> Path:
    if getattr(sys, "frozen", False):
        base = Path(os.environ.get("APPDATA", Path.home())) / "MAX POST"
    else:
        base = Path(__file__).parent
    base.mkdir(parents=True, exist_ok=True)
    return base / "stats_cache.json"


def _save_cache(rows: list[dict], summary_texts: list[str],
                group_cache: dict | None = None) -> None:
    path = _cache_path()
    data = {
        "ts":           datetime.now().isoformat(),
        "rows":         rows,
        "summary_texts": summary_texts,
        "group_cache":  group_cache or {},
    }
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", encoding="utf-8", dir=path.parent, delete=False, suffix=".tmp"
        ) as tmp:
            json.dump(data, tmp, ensure_ascii=False, indent=2)
            tmp_path = Path(tmp.name)
        tmp_path.replace(path)
    except Exception as exc:
        _log.warning("stats cache write error: %s", exc)


def _load_cache() -> tuple[list[dict], list[str], datetime | None, dict]:
    """Возвращает (rows, summary_texts, ts, group_cache).

    group_cache: {chat_id: {"name": str, "members": str, "cached_at": int}}
    """
    try:
        raw = json.loads(_cache_path().read_text(encoding="utf-8"))
        ts = datetime.fromisoformat(raw["ts"])
        return raw["rows"], raw.get("summary_texts", []), ts, raw.get("group_cache", {})
    except Exception:
        return [], [], None, {}


# ────────────────────────────────────────────────────────────────
#  Фоновый поток — GREEN-API
# ────────────────────────────────────────────────────────────────

class _FetchWorker(QThread):
    data_ready = pyqtSignal(list, list, dict)  # (rows, summary_texts, group_cache) — финал
    row_ready  = pyqtSignal(dict)              # одна строка — сразу в таблицу
    failed     = pyqtSignal(str)
    progress   = pyqtSignal(str)               # текст для строки статуса

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._stop = False

    def run(self) -> None:
        load_dotenv(get_env_path())
        api_url   = os.getenv("MAX_API_URL", "https://api.green-api.com")
        id_inst   = os.getenv("MAX_ID_INSTANCE", "")
        api_token = os.getenv("MAX_API_TOKEN", "")

        if not id_inst or not api_token:
            self.failed.emit(
                "Не заданы MAX_ID_INSTANCE или MAX_API_TOKEN.\n"
                "Откройте Настройки подключений (🔑) и заполните данные GREEN-API."
            )
            return

        # ── 1. Читаем chat_id из Excel ───────────────────────────
        excel_path = _resolve_excel_path()
        if not excel_path.exists():
            self.failed.emit(f"Файл {excel_path.name} не найден рядом с программой.")
            return

        try:
            df = pd.read_excel(excel_path, dtype=str).fillna("")
        except Exception as exc:
            self.failed.emit(f"Ошибка чтения {excel_path.name}: {exc}")
            return

        cols     = {str(c).strip().lower(): c for c in df.columns}
        addr_col = cols.get("адрес") or df.columns[0]
        link_col = cols.get("ссылка") or (df.columns[1] if len(df.columns) > 1 else None)
        id_col   = cols.get("id")    or (df.columns[2] if len(df.columns) > 2 else None)

        entries: list[tuple[str, str, str]] = []   # (chat_id, link, address)
        for _, row in df.iterrows():
            raw_id  = str(row.get(id_col, "")).strip() if id_col else ""
            link    = str(row.get(link_col, "")).strip() if link_col else ""
            address = str(row.get(addr_col, "")).strip()

            # Убираем .0 (числовые ID из Excel читаются как float)
            if raw_id.endswith(".0"):
                raw_id = raw_id[:-2]

            # Пробуем получить числовой ID: из колонки ID или из ссылки
            chat_id = _extract_chat_id(raw_id) if raw_id and raw_id.lower() not in ("nan", "") \
                      else _extract_chat_id(link)

            if not chat_id or chat_id.lower() in ("nan", "") or not chat_id.lstrip("-").isdigit():
                continue
            entries.append((chat_id, link, address))

        if not entries:
            self.failed.emit(f"В файле {excel_path.name} не найдены ID групп (колонка «ID»).")
            return

        # ── 2. Загружаем данные групп (умный кэш + 1 req/s) ─────
        _, _, _, group_cache = _load_cache()
        now_ts  = int(time.time())
        rows: list[dict] = []
        total   = len(entries)

        # Удаляем из кэша записи групп, которых нет в текущем Excel
        current_ids = {cid for cid, _, _ in entries}
        stale_keys  = [k for k in group_cache if k not in current_ids]
        for k in stale_keys:
            del group_cache[k]

        # Считаем сколько запросов нужно (не в кэше или кэш устарел)
        need_fetch = sum(
            1 for cid, _, _ in entries
            if now_ts - group_cache.get(cid, {}).get("cached_at", 0) > _GROUP_CACHE_TTL
        )
        need_activity = sum(
            1 for cid, _, _ in entries
            if now_ts - group_cache.get(cid, {}).get("activity_cached_at", 0) > _ACTIVITY_CACHE_TTL
        )
        total_requests = need_fetch + need_activity
        from_cache = total - need_fetch
        if from_cache > 0:
            est_min = round(total_requests * _REQUEST_DELAY / 60, 1)
            self.progress.emit(
                f"Загрузка {need_fetch} групп (~{est_min} мин), "
                f"{from_cache} из кэша…  0 / {total_requests}"
            )
        else:
            est_min = round(total_requests * _REQUEST_DELAY / 60, 1)
            self.progress.emit(f"Загрузка {total} групп (~{est_min} мин)…  0 / {total_requests}")

        fetched = 0   # счётчик реальных запросов (не кэш)
        first_request = True
        url_group    = f"{api_url}/waInstance{id_inst}/getGroupData/{api_token}"
        url_history  = f"{api_url}/waInstance{id_inst}/getChatHistory/{api_token}"

        for chat_id, link, address in entries:
            if self._stop:
                return
            cached = group_cache.get(chat_id, {})
            cache_age = now_ts - cached.get("cached_at", 0)

            if cache_age <= _GROUP_CACHE_TTL and cached.get("name"):
                # ── Берём из кэша ──────────────────────────────────
                name    = cached["name"]
                members = cached["members"]
            else:
                # ── Запрашиваем у GREEN-API ────────────────────────
                if not first_request:
                    time.sleep(_REQUEST_DELAY)   # 1 req/s лимит GREEN-API
                first_request = False
                fetched += 1
                self.progress.emit(
                    f"Загрузка групп… {fetched} / {total_requests}"
                    + (f"  (кэш: {from_cache})" if from_cache > 0 else "")
                )

                name    = address
                members = "—"   # нет доступа по умолчанию
                for attempt in range(2):
                    try:
                        resp = requests.post(
                            url_group, json={"chatId": chat_id}, timeout=20
                        )
                        if resp.ok:
                            data    = resp.json()
                            name    = data.get("subject") or address
                            members = str(data.get("size", 0))
                            group_cache[chat_id] = {
                                **group_cache.get(chat_id, {}),
                                "name":      name,
                                "members":   members,
                                "cached_at": now_ts,
                            }
                        else:
                            members = "—"   # API вернул ошибку — реально нет доступа
                        break  # выходим из retry-цикла (успех или нет доступа)
                    except requests.exceptions.Timeout:
                        _log.warning("getGroupData %s: timeout (попытка %d)", chat_id, attempt + 1)
                        if attempt == 0:
                            time.sleep(3)   # ждём 3с перед повтором
                            continue
                        members = "~"   # таймаут — неизвестно, доступ может быть
                    except Exception as exc:
                        _log.warning("getGroupData %s: %s", chat_id, exc)
                        members = "~"   # сетевая ошибка
                        break

            # ── getChatHistory: последнее сообщение (кэш 6 ч) ────
            cached = group_cache.get(chat_id, {})
            activity_age = now_ts - cached.get("activity_cached_at", 0)
            if members not in ("—",) and activity_age > _ACTIVITY_CACHE_TTL:
                if not first_request:
                    time.sleep(_REQUEST_DELAY)
                first_request = False
                fetched += 1
                self.progress.emit(
                    f"Загрузка активности… {fetched} / {total_requests}"
                    + (f"  (кэш: {from_cache})" if from_cache > 0 else "")
                )
                try:
                    resp = requests.post(
                        url_history,
                        json={"chatId": chat_id, "count": 1},
                        timeout=12,
                    )
                    if resp.ok:
                        history = resp.json()
                        ts_val = history[0].get("timestamp", 0) if isinstance(history, list) and history else 0
                        group_cache.setdefault(chat_id, {})["last_activity"] = ts_val
                        group_cache[chat_id]["activity_cached_at"] = now_ts
                        _log.debug("getChatHistory %s: ts=%s", chat_id, ts_val)
                    else:
                        _log.debug("getChatHistory %s: HTTP %d", chat_id, resp.status_code)
                        group_cache.setdefault(chat_id, {})["activity_cached_at"] = now_ts
                except Exception as exc:
                    _log.warning("getChatHistory %s: %s", chat_id, exc)
                    # Кэшируем факт ошибки — не повторяем до истечения TTL
                    group_cache.setdefault(chat_id, {})["activity_cached_at"] = now_ts

            ts         = group_cache.get(chat_id, {}).get("last_activity")
            last_event = (
                datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M")
                if ts else "Нет данных"
            )
            if members == "—":
                last_event = "Нет доступа"
            elif members == "~":
                last_event = "Ошибка сети"

            row = {
                "name":       name,
                "members":    members,
                "last_event": last_event,
                "link":       link,
                "chat_id":    chat_id,
            }
            rows.append(row)
            self.row_ready.emit(row)

        if not rows:
            self.failed.emit("Не удалось получить данные ни по одной группе.")
            return

        self.data_ready.emit(rows, [], group_cache)


# ────────────────────────────────────────────────────────────────
#  Быстрый воркер: парсим web-отчёт (1 запрос вместо ~190)
# ────────────────────────────────────────────────────────────────

class _WebFetchWorker(QThread):
    done     = pyqtSignal(list)   # список строк
    failed   = pyqtSignal(str)
    progress = pyqtSignal(str)

    def run(self) -> None:
        self.progress.emit("Загрузка отчёта с сервера…")
        try:
            from bs4 import BeautifulSoup
            resp = requests.get(_WEB_REPORT_URL, timeout=15)
            resp.raise_for_status()
            soup  = BeautifulSoup(resp.text, "html.parser")
            table = soup.find("table")
            if not table:
                self.failed.emit("Таблица не найдена на странице отчёта.")
                return
            rows = []
            for tr in table.find_all("tr")[1:]:
                tds = tr.find_all("td")
                if len(tds) < 3:
                    continue
                name    = tds[0].get_text(strip=True)
                members = tds[1].get_text(strip=True)
                ts_str  = tds[2].get_text(strip=True)
                link    = ""
                if len(tds) >= 4:
                    a = tds[3].find("a")
                    link = a["href"] if a else tds[3].get_text(strip=True)
                rows.append({
                    "name":       name,
                    "members":    members,
                    "last_event": ts_str,
                    "link":       link,
                    "chat_id":    "",
                })
            self.done.emit(rows)
        except Exception as exc:
            self.failed.emit(str(exc))


# ────────────────────────────────────────────────────────────────
#  Воркер: запрос отчёта у бота через GREEN-API
# ────────────────────────────────────────────────────────────────

class _BotReportWorker(QThread):
    """Отправляет команду боту и скачивает HTML-файл ответа."""
    progress = pyqtSignal(str)
    done     = pyqtSignal(str)   # HTML-содержимое файла
    failed   = pyqtSignal(str)

    _BOT_CHAT_ID  = "69347387@c.us"
    _POLL_SEC     = 4
    _TIMEOUT_SEC  = 120

    def __init__(self, instance_id: str, token: str, days: int, parent=None):
        super().__init__(parent)
        self._instance_id = instance_id
        self._token       = token
        self._days        = days
        self._stop        = False

    def stop(self) -> None:
        self._stop = True

    def run(self) -> None:
        base    = f"https://api.green-api.com/waInstance{self._instance_id}"
        command = f"/мах_отчет_история {self._days}"

        # ── 1. Отправляем команду ────────────────────────────────
        self.progress.emit("Отправка команды боту…")
        sent_at = int(time.time())
        try:
            r = requests.post(
                f"{base}/sendMessage/{self._token}",
                json={"chatId": self._BOT_CHAT_ID, "message": command},
                timeout=15,
            )
            r.raise_for_status()
        except Exception as e:
            self.failed.emit(f"Ошибка отправки команды боту: {e}")
            return

        # ── 2. Ждём HTML-файл в ответе бота ─────────────────────
        deadline = time.time() + self._TIMEOUT_SEC
        tick     = 0
        while not self._stop and time.time() < deadline:
            time.sleep(self._POLL_SEC)
            tick += self._POLL_SEC
            self.progress.emit(
                f"Ожидание ответа бота… {tick} / {self._TIMEOUT_SEC} сек"
            )
            try:
                r = requests.post(
                    f"{base}/getChatHistory/{self._token}",
                    json={"chatId": self._BOT_CHAT_ID, "count": 10},
                    timeout=20,
                )
                r.raise_for_status()
                messages = r.json()
            except Exception:
                continue

            for msg in messages:
                if msg.get("timestamp", 0) <= sent_at:
                    continue
                if msg.get("typeMessage") != "documentMessage":
                    continue
                fd    = msg.get("fileMessageData", {})
                fname = fd.get("fileName", "")
                mime  = fd.get("mimeType", "")
                dl    = fd.get("downloadUrl", "")
                if not dl:
                    continue
                if not (fname.lower().endswith(".html") or "html" in mime):
                    continue
                self.progress.emit("Скачивание файла отчёта…")
                try:
                    r2 = requests.get(dl, timeout=30)
                    r2.raise_for_status()
                    self.done.emit(r2.text)
                except Exception as e:
                    self.failed.emit(f"Ошибка скачивания файла: {e}")
                return

        if not self._stop:
            self.failed.emit(
                f"Бот не ответил за {self._TIMEOUT_SEC} сек.\n"
                "Попробуйте ещё раз или загрузите файл вручную."
            )


# ────────────────────────────────────────────────────────────────
#  Виджет панели
# ────────────────────────────────────────────────────────────────

class StatsPanel(QWidget):
    """Панель «Статистика групп» — таблица с реалтайм-данными из GREEN-API."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._worker:     _FetchWorker | _WebFetchWorker | None = None
        self._bot_worker: _BotReportWorker | None = None
        self._all_rows: list[dict] = []
        self._missing_rows: list[dict] = []
        self._dead_only: bool = False
        self._history_data: dict   = {}   # name_key → {delta, latest, oldest}
        self._history_period: str  = ""
        self._period_days: int = 0
        self._last_refresh: datetime | None = None

        self._spin_frame: int = 0
        self._spin_timer = QTimer(self)
        self._spin_timer.setInterval(100)
        self._spin_timer.timeout.connect(self._spin_step)

        self._build_ui()
        self._apply_styles()

        # Авто-обновление каждые 5 минут
        self._auto_timer = QTimer(self)
        self._auto_timer.setInterval(_AUTO_REFRESH_MS)
        self._auto_timer.timeout.connect(self.refresh)
        self._auto_timer.start()

        QApplication.instance().aboutToQuit.connect(self._shutdown)

        # Первая загрузка
        QTimer.singleShot(0, self.refresh)

    # ── UI ──────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        # ── Переключатель режимов (pill) ─────────────────────────
        toggle_row = QHBoxLayout()
        toggle_row.setSpacing(0)
        self._btn_web   = QPushButton("🌐  WEB версия")
        self._btn_smart = QPushButton("⚡  Smart версия")
        for btn in (self._btn_web, self._btn_smart):
            btn.setCheckable(True)
            btn.setFixedHeight(34)
            btn.setObjectName("statsToggleBtn")
        self._btn_smart.setObjectName("statsToggleBtnActive")
        self._btn_web.setChecked(False)
        self._btn_smart.setChecked(True)
        self._btn_web.clicked.connect(lambda: self._switch_mode(0))
        self._btn_smart.clicked.connect(lambda: self._switch_mode(1))
        toggle_row.addStretch()
        toggle_row.addWidget(self._btn_web)
        toggle_row.addWidget(self._btn_smart)
        toggle_row.addStretch()
        root.addLayout(toggle_row)

        # ── Стек страниц ─────────────────────────────────────────
        self._mode_stack = QStackedWidget()
        root.addWidget(self._mode_stack)

        # ── Страница 0: WEB версия ────────────────────────────────
        if _WEB_ENGINE_AVAILABLE:
            self._web_view = QWebEngineView()
            _page = QWebEnginePage(self._web_view)
            _page.certificateError.connect(lambda err: err.acceptCertificate())
            self._web_view.setPage(_page)
            self._web_view.loadFinished.connect(self._on_web_load_finished)
            self._web_view.setUrl(QUrl(_WEB_REPORT_URL))
            self._mode_stack.addWidget(self._web_view)
        else:
            no_web = QLabel(
                "⚠️  Для WEB версии требуется пакет PyQt6-WebEngine.\n"
                "Установите его командой: pip install PyQt6-WebEngine"
            )
            no_web.setAlignment(Qt.AlignmentFlag.AlignCenter)
            no_web.setObjectName("statsStatus")
            self._mode_stack.addWidget(no_web)

        # ── Страница 1: Smart версия ──────────────────────────────
        smart_page = QWidget()
        smart_layout = QVBoxLayout(smart_page)
        smart_layout.setContentsMargins(0, 0, 0, 0)
        smart_layout.setSpacing(10)
        self._mode_stack.addWidget(smart_page)
        self._mode_stack.setCurrentIndex(1)  # по умолчанию Smart версия

        # ── Заголовок ────────────────────────────────────────────
        hdr = QHBoxLayout()
        title = QLabel("Статистика групп")
        title.setObjectName("statsPanelTitle")
        hdr.addWidget(title)
        hdr.addStretch()

        self._last_lbl = QLabel("")
        self._last_lbl.setObjectName("statsLastRefresh")
        hdr.addWidget(self._last_lbl)

        self._export_btn = QPushButton("📥 Excel")
        self._export_btn.setObjectName("statsExportBtn")
        self._export_btn.setFixedHeight(32)
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_excel)
        hdr.addWidget(self._export_btn)

        self._history_btn = QPushButton("📊 История")
        self._history_btn.setObjectName("statsHistoryBtn")
        self._history_btn.setFixedHeight(32)
        self._history_btn.setToolTip("Загрузить HTML-отчёт «История подписчиков»")
        self._history_btn.clicked.connect(self._load_subscriber_history)
        hdr.addWidget(self._history_btn)

        self._refresh_btn = QPushButton("⟳  Обновить")
        self._refresh_btn.setObjectName("statsRefreshBtn")
        self._refresh_btn.setFixedHeight(32)
        self._refresh_btn.clicked.connect(self.refresh)
        hdr.addWidget(self._refresh_btn)
        smart_layout.addLayout(hdr)

        # ── Сводка ───────────────────────────────────────────────
        self._summary_frame = QFrame()
        self._summary_frame.setObjectName("statsSummary")
        sf_layout = QHBoxLayout(self._summary_frame)
        sf_layout.setContentsMargins(14, 8, 14, 8)
        sf_layout.setSpacing(28)

        self._lbl_groups    = self._make_stat_lbl("—", "Групп")
        self._lbl_members   = self._make_stat_lbl("—", "Участников")
        self._lbl_active    = self._make_stat_lbl("—", "Активны сегодня")
        self._lbl_yesterday = self._make_stat_lbl("—", "Активны вчера")
        self._lbl_missing   = self._make_stat_lbl("—", "Нет доступа")

        for w in (self._lbl_groups, self._lbl_members, self._lbl_active,
                  self._lbl_yesterday, self._lbl_missing):
            sf_layout.addWidget(w)
        sf_layout.addStretch()
        smart_layout.addWidget(self._summary_frame)

        # ── Баннер кэша (скрыт по умолчанию) ────────────────────
        self._cache_banner = QLabel()
        self._cache_banner.setObjectName("statsCacheBanner")
        self._cache_banner.setWordWrap(True)
        self._cache_banner.hide()
        smart_layout.addWidget(self._cache_banner)

        self._history_banner = QLabel()
        self._history_banner.setObjectName("statsHistoryBanner")
        self._history_banner.setWordWrap(True)
        self._history_banner.hide()
        smart_layout.addWidget(self._history_banner)

        # ── Прогресс-бар загрузки (скрыт по умолчанию) ──────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setObjectName("statsProgressBar")
        self._progress_bar.setFixedHeight(6)
        self._progress_bar.setRange(0, 100)
        self._progress_bar.setValue(0)
        self._progress_bar.setTextVisible(False)
        self._progress_bar.hide()
        smart_layout.addWidget(self._progress_bar)

        # ── Фильтр по периоду ────────────────────────────────────
        period_row = QHBoxLayout()
        period_row.setSpacing(4)
        self._period_btns: dict[int, QPushButton] = {}
        for days, label in ((0, "Все"), (1, "День"), (7, "Неделя"), (30, "Месяц")):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(28)
            btn.setObjectName("statsPeriodBtnActive" if days == 0 else "statsPeriodBtn")
            btn.setChecked(days == 0)
            btn.clicked.connect(lambda _, d=days: self._switch_period(d))
            period_row.addWidget(btn)
            self._period_btns[days] = btn
        period_row.addStretch()
        smart_layout.addLayout(period_row)

        # ── Сводка периода ───────────────────────────────────────
        self._period_summary = QLabel("")
        self._period_summary.setObjectName("statsPeriodSummary")
        self._period_summary.setWordWrap(True)
        self._period_summary.hide()
        smart_layout.addWidget(self._period_summary)

        # ── Поиск ────────────────────────────────────────────────
        search_row = QHBoxLayout()
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍  Поиск по названию...")
        self._search.setObjectName("statsSearch")
        self._search.textChanged.connect(self._apply_filter)
        search_row.addWidget(self._search)

        self._dead_btn = QPushButton("🔴 Нет доступа")
        self._dead_btn.setObjectName("statsDeadBtn")
        self._dead_btn.setCheckable(True)
        self._dead_btn.setEnabled(False)
        self._dead_btn.toggled.connect(self._on_dead_toggled)
        search_row.addWidget(self._dead_btn)
        smart_layout.addLayout(search_row)

        # ── Таблица ───────────────────────────────────────────────
        self._table = QTableWidget(0, 5)
        self._table.setObjectName("statsTable")
        self._table.setHorizontalHeaderLabels(
            ["Название", "Участников", "Последняя активность", "Δ Подписчики", "Ссылка"]
        )
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(1, 110)
        self._table.setColumnWidth(2, 180)
        self._table.setColumnWidth(3, 110)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(True)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSortIndicatorShown(True)
        self._table.cellDoubleClicked.connect(self._on_double_click)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._show_context_menu)
        smart_layout.addWidget(self._table)

        # ── Статус-строка ─────────────────────────────────────────
        self._status_lbl = QLabel("Загрузка…")
        self._status_lbl.setObjectName("statsStatus")
        smart_layout.addWidget(self._status_lbl)


    def _switch_mode(self, index: int) -> None:
        self._mode_stack.setCurrentIndex(index)
        self._btn_web.setChecked(index == 0)
        self._btn_smart.setChecked(index == 1)
        self._btn_web.setObjectName("statsToggleBtnActive" if index == 0 else "statsToggleBtn")
        self._btn_smart.setObjectName("statsToggleBtnActive" if index == 1 else "statsToggleBtn")
        for btn in (self._btn_web, self._btn_smart):
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def _on_web_load_finished(self, ok: bool) -> None:
        if not ok and _WEB_ENGINE_AVAILABLE and hasattr(self, "_web_view"):
            self._web_view.setHtml("""
                <html><body style="font-family:Arial,sans-serif;text-align:center;
                       padding:80px 40px;color:#6b7280;background:#f9fafb;">
                  <div style="font-size:48px;margin-bottom:16px">⚠️</div>
                  <h2 style="color:#374151;margin:0 0 8px">Сервер недоступен</h2>
                  <p style="margin:0 0 24px">Не удалось загрузить WEB версию отчёта.</p>
                  <p style="font-size:13px;background:#e5e7eb;padding:10px 20px;
                     border-radius:8px;display:inline-block">
                    Используйте <b>⚡ Smart версию</b> — она работает напрямую через GREEN-API
                  </p>
                </body></html>
            """)

    @staticmethod
    def _make_stat_lbl(value: str, label: str) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(1)
        val_lbl = QLabel(value)
        val_lbl.setObjectName("statsStatValue")
        lbl_lbl = QLabel(label)
        lbl_lbl.setObjectName("statsStatLabel")
        lay.addWidget(val_lbl, alignment=Qt.AlignmentFlag.AlignHCenter)
        lay.addWidget(lbl_lbl, alignment=Qt.AlignmentFlag.AlignHCenter)
        w._val_lbl = val_lbl  # type: ignore[attr-defined]
        return w

    # ── Логика обновления ────────────────────────────────────────

    _SPINNER = ("⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏")

    def _spin_step(self) -> None:
        self._spin_frame = (self._spin_frame + 1) % len(self._SPINNER)
        self._refresh_btn.setText(f"{self._SPINNER[self._spin_frame]}  Загрузка…")

    def refresh(self) -> None:
        if self._worker and self._worker.isRunning():
            return
        self._refresh_btn.setEnabled(False)
        self._spin_frame = 0
        self._spin_timer.start()
        self._refresh_btn.setText(f"{self._SPINNER[0]}  Загрузка…")

        # Сразу показываем кэш, чтобы пользователь видел данные во время загрузки
        cached_rows, _, cached_ts, _ = _load_cache()
        if cached_rows and not self._all_rows:
            self._all_rows = cached_rows
            self._apply_filter()
            ts_str = cached_ts.strftime("%d.%m.%Y %H:%M") if cached_ts else "ранее"
            self._cache_banner.setText(f"🔄  Обновление данных… показаны данные от {ts_str}")
            self._cache_banner.show()
        else:
            self._cache_banner.hide()

        self._progress_bar.setValue(0)
        self._progress_bar.show()
        self._status_lbl.setText("Загрузка отчёта с сервера…")

        self._worker = _WebFetchWorker(self)
        self._worker.done.connect(self._on_web_data)
        self._worker.failed.connect(self._on_error)
        self._worker.progress.connect(self._on_progress)
        self._worker.done.connect(self._worker.deleteLater)
        self._worker.failed.connect(self._worker.deleteLater)
        self._worker.done.connect(lambda *_: setattr(self, "_worker", None))
        self._worker.failed.connect(lambda *_: setattr(self, "_worker", None))
        self._worker.start()

    def _on_progress(self, text: str) -> None:
        self._status_lbl.setText(text)
        # Парсим "Загрузка групп… X / Y" для прогресс-бара
        if "/" in text:
            try:
                parts = text.split("…")[-1].strip().split("/")
                current = int(parts[0].strip().split()[0])
                total   = int(parts[1].strip().split()[0])
                if total > 0:
                    self._progress_bar.setValue(int(current / total * 100))
            except (ValueError, IndexError):
                pass

    def _on_row_ready(self, row: dict) -> None:
        """Добавляет одну строку в таблицу сразу по мере загрузки."""
        if not self._streaming_started:
            # Первая живая строка — очищаем кэш-отображение и начинаем свежий список
            self._streaming_started = True
            self._all_rows = []
            self._missing_rows = []
            self._table.setSortingEnabled(False)
            self._table.setRowCount(0)
            self._cache_banner.hide()

        self._all_rows.append(row)
        if row["members"] in ("—", "~"):
            self._missing_rows.append(row)

        # Пропускаем если активен фильтр поиска и строка не подходит
        query = self._search.text().strip().lower()
        if self._dead_only:
            if row["members"] != "—":
                return
            if query and query not in row["name"].lower():
                return
        elif query and query not in row["name"].lower():
            return

        self._append_table_row(row)
        self._update_summary_labels()

        total = len(self._all_rows)
        shown = self._table.rowCount()
        self._status_lbl.setText(f"Загружено {total} групп…")
        if query or self._dead_only:
            self._status_lbl.setText(f"Показано {shown} из {total} групп…")

    def _append_table_row(self, row: dict) -> None:
        """Добавляет одну строку в конец таблицы."""
        today = datetime.now().date()
        t_event = row["last_event"]
        members = row["members"]

        fresh_today = False
        fresh_week  = False
        try:
            ev_dt = datetime.strptime(t_event[:10], "%d.%m.%Y").date()
            fresh_today = (ev_dt == today)
            fresh_week  = (ev_dt >= today - timedelta(days=7))
        except ValueError:
            pass

        if members == "—":
            row_bg = QColor("#fff0f0")   # красный — нет доступа
        elif members == "~":
            row_bg = QColor("#fffbee")   # жёлтый — ошибка сети/таймаут
        elif fresh_today:
            row_bg = QColor("#f0faf2")   # зелёный — активна сегодня
        elif fresh_week:
            row_bg = QColor("#f0f7ff")   # голубой — активна на этой неделе
        else:
            row_bg = None

        # Отключаем сортировку на время вставки — иначе новые строки
        # могут попасть в случайное место при включённом sort indicator
        was_sorting = self._table.isSortingEnabled()
        self._table.setSortingEnabled(False)

        row_idx = self._table.rowCount()
        self._table.insertRow(row_idx)
        name_item = QTableWidgetItem(row["name"])
        name_item.setData(Qt.ItemDataRole.UserRole, row.get("chat_id", ""))
        delta_item = self._make_delta_item(row["name"])
        items = [
            name_item,
            _NumItem(members),
            QTableWidgetItem(t_event),
            delta_item,
            QTableWidgetItem(row["link"]),
        ]
        for col, item in enumerate(items):
            item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            if row_bg:
                item.setBackground(row_bg)
            self._table.setItem(row_idx, col, item)

        self._table.setSortingEnabled(was_sorting)

    @staticmethod
    def _history_key(name: str) -> str:
        """Нормализованный ключ для сопоставления с историей подписчиков."""
        import unicodedata
        s = unicodedata.normalize("NFC", name.strip().lower())
        return " ".join(s.split())

    def _make_delta_item(self, name: str) -> QTableWidgetItem:
        """Создаёт ячейку Δ Подписчики для строки таблицы."""
        key  = self._history_key(name)
        info = self._history_data.get(key)
        if info is None:
            item = _NumItem("—")
            item.setForeground(QColor("#9ca3af"))
            return item
        delta = info["delta"]
        if delta > 0:
            text = f"+{delta}"
            color = QColor("#22c55e")
        elif delta < 0:
            text = str(delta)
            color = QColor("#ef4444")
        else:
            text = "="
            color = QColor("#9ca3af")
        item = _NumItem(text)
        item.setForeground(color)
        font = QFont()
        font.setBold(True)
        item.setFont(font)
        return item

    def _update_summary_labels(self) -> None:
        """Обновляет цифры сводки по текущему _all_rows."""
        rows = self._all_rows
        total = len(rows)
        total_members = 0
        active_today     = 0
        active_yesterday = 0
        today     = datetime.now().date()
        yesterday = today - timedelta(days=1)
        for r in rows:
            try:
                total_members += int(r["members"])
            except (ValueError, TypeError):
                pass
            try:
                ev_dt = datetime.strptime(r["last_event"][:10], "%d.%m.%Y").date()
                if ev_dt == today:
                    active_today += 1
                elif ev_dt == yesterday:
                    active_yesterday += 1
            except ValueError:
                pass
        missing_count = len(self._missing_rows)

        self._lbl_groups._val_lbl.setText(str(total))                                    # type: ignore[attr-defined]
        self._lbl_members._val_lbl.setText(f"{total_members:,}".replace(",", " "))       # type: ignore[attr-defined]
        self._lbl_active._val_lbl.setText(str(active_today))                             # type: ignore[attr-defined]
        self._lbl_yesterday._val_lbl.setText(str(active_yesterday))                      # type: ignore[attr-defined]
        missing_val_lbl = self._lbl_missing._val_lbl                                     # type: ignore[attr-defined]
        missing_val_lbl.setText(str(missing_count))
        missing_val_lbl.setStyleSheet(
            "color: #dc2626; font-size: 24px; font-weight: 700;"
            if missing_count > 0 else ""
        )

    def _on_web_data(self, rows: list[dict]) -> None:
        self._spin_timer.stop()
        self._progress_bar.setValue(100)
        self._progress_bar.hide()
        self._all_rows    = rows
        self._missing_rows = []
        self._last_refresh = datetime.now()
        self._last_lbl.setText(f"Обновлено в {self._last_refresh.strftime('%H:%M:%S')}")
        self._cache_banner.hide()
        _save_cache(rows, [], {})
        self._refresh_btn.setEnabled(True)
        self._refresh_btn.setText("⟳  Обновить")
        self._export_btn.setEnabled(True)
        self._dead_btn.setEnabled(False)
        self._table.setSortingEnabled(True)
        self._apply_filter()
        self._update_summary_labels()
        total = len(rows)
        self._status_lbl.setText(
            f"Загружено {total} групп · "
            f"обновлено {self._last_refresh.strftime('%d.%m.%Y %H:%M:%S')}"
        )

    def _switch_period(self, days: int) -> None:
        self._period_days = days
        for d, btn in self._period_btns.items():
            active = (d == days)
            btn.setChecked(active)
            btn.setObjectName("statsPeriodBtnActive" if active else "statsPeriodBtn")
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        self._apply_filter()

    _PERIOD_LABELS = {0: None, 1: "сегодня", 7: "за неделю", 30: "за месяц"}

    def _update_period_summary(self, filtered_rows: list[dict]) -> None:
        label = self._PERIOD_LABELS.get(self._period_days)
        if not label or not self._all_rows:
            self._period_summary.hide()
            return
        total_all   = len(self._all_rows)
        active      = len(filtered_rows)
        members     = sum(int(r["members"]) for r in filtered_rows
                         if str(r["members"]).isdigit())
        members_all = sum(int(r["members"]) for r in self._all_rows
                         if str(r["members"]).isdigit())
        self._period_summary.setText(
            f"📊  {label.capitalize()}: активны <b>{active}</b> из {total_all} групп  ·  "
            f"участников в активных: <b>{members:,}</b> из {members_all:,}"
            .replace(",", "\u00a0")
        )
        self._period_summary.show()

    def _on_data(self, rows: list[dict], summary_texts: list[str], group_cache: dict) -> None:
        self._spin_timer.stop()
        self._progress_bar.setValue(100)
        self._progress_bar.hide()
        # rows уже добавлены в _all_rows через row_ready — синхронизируем на всякий случай
        if rows and not self._all_rows:
            self._all_rows = rows
            self._missing_rows = [r for r in rows if r["members"] in ("—", "~")]
        self._last_refresh = datetime.now()
        self._last_lbl.setText(f"Обновлено в {self._last_refresh.strftime('%H:%M:%S')}")
        self._cache_banner.hide()
        _save_cache(rows, summary_texts, group_cache)
        self._refresh_btn.setEnabled(True)
        self._refresh_btn.setText("⟳  Обновить")
        self._export_btn.setEnabled(True)
        self._dead_btn.setEnabled(True)
        # Включаем сортировку (была отключена во время стриминга)
        self._table.setSortingEnabled(True)
        self._update_summary_labels()
        total = len(self._all_rows)
        self._status_lbl.setText(
            f"Загружено {total} групп · "
            f"обновлено {self._last_refresh.strftime('%d.%m.%Y %H:%M:%S')}"
        )

    def _on_error(self, msg: str) -> None:
        self._spin_timer.stop()
        self._progress_bar.hide()
        self._refresh_btn.setEnabled(True)
        self._refresh_btn.setText("⟳  Обновить")
        self._table.setSortingEnabled(True)
        cached_rows, cached_summary, cached_ts, _ = _load_cache()
        if cached_rows:
            self._all_rows = cached_rows
            self._missing_rows = [r for r in cached_rows if r["members"] in ("—", "~")]
            ts_str = cached_ts.strftime("%d.%m.%Y %H:%M") if cached_ts else "неизвестно"
            self._cache_banner.setText(
                f"⚠️  Ошибка загрузки: {msg}  ·  Показаны данные от {ts_str}"
            )
            self._cache_banner.show()
            self._export_btn.setEnabled(True)
            self._apply_filter()
            self._status_lbl.setText(f"Кэш от {ts_str} · {len(cached_rows)} групп")
        else:
            self._cache_banner.hide()
            self._status_lbl.setText(f"Ошибка: {msg}  ·  Кэш недоступен")

    def _on_dead_toggled(self, checked: bool) -> None:
        self._dead_only = checked
        self._apply_filter()

    def _apply_filter(self) -> None:
        query = self._search.text().strip().lower()

        # Фильтр по периоду
        if self._period_days > 0:
            cutoff = datetime.now() - timedelta(days=self._period_days)
            base = []
            for r in self._all_rows:
                try:
                    if datetime.strptime(r["last_event"][:10], "%d.%m.%Y") >= cutoff:
                        base.append(r)
                except ValueError:
                    pass
        else:
            base = self._all_rows

        if self._dead_only:
            rows = [r for r in self._missing_rows if not query or query in r["name"].lower()]
            self._fill_table(rows, dead_mode=True)
            shown = len(rows)
            total = len(self._missing_rows)
            self._status_lbl.setText(
                f"Нет доступа: показано {shown} из {total}"
                + (f" · фильтр: «{query}»" if query else "")
            )
        else:
            rows = [r for r in base if not query or query in r["name"].lower()]
            self._fill_table(rows)
            self._update_period_summary(base)
            if self._all_rows:
                shown = len(rows)
                total = len(self._all_rows)
                self._status_lbl.setText(
                    f"Показано {shown} из {total} групп"
                    + (f" · фильтр: «{query}»" if query else "")
                )

    def _fill_table(self, rows: list[dict], *, dead_mode: bool = False) -> None:
        today = datetime.now().date()
        self._table.setSortingEnabled(False)
        self._table.setRowCount(len(rows))

        dead_bg = QColor("#fff0f0")

        for row_idx, r in enumerate(rows):
            name    = r["name"]
            members = r["members"]
            t_event = r["last_event"]
            link    = r["link"]

            if dead_mode:
                row_bg = dead_bg
            elif members == "~":
                row_bg = QColor("#fffbee")   # жёлтый — ошибка сети
            else:
                fresh_today = False
                fresh_week  = False
                try:
                    ev_dt = datetime.strptime(t_event[:10], "%d.%m.%Y").date()
                    fresh_today = (ev_dt == today)
                    fresh_week  = (ev_dt >= today - timedelta(days=7))
                except ValueError:
                    pass
                if fresh_today:
                    row_bg = QColor("#f0faf2")
                elif fresh_week:
                    row_bg = QColor("#f0f7ff")
                else:
                    row_bg = None

            name_item = QTableWidgetItem(name)
            name_item.setData(Qt.ItemDataRole.UserRole, r.get("chat_id", ""))
            delta_item = self._make_delta_item(name)
            items: list[QTableWidgetItem] = [
                name_item,
                _NumItem(members),
                QTableWidgetItem(t_event),
                delta_item,
                QTableWidgetItem(link),
            ]
            for col, item in enumerate(items):
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if row_bg:
                    item.setBackground(row_bg)
                self._table.setItem(row_idx, col, item)

        self._table.setSortingEnabled(True)
        self._table.resizeRowsToContents()

    # ── Взаимодействие с таблицей ────────────────────────────────

    def _get_row_link(self, row: int) -> str:
        item = self._table.item(row, _COL_LINK)
        return item.text().strip() if item else ""

    def _on_double_click(self, row: int, _col: int) -> None:
        link = self._get_row_link(row)
        if link.startswith("http"):
            QDesktopServices.openUrl(QUrl(link))

    def _show_context_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if row < 0:
            return
        link = self._get_row_link(row)
        name_item = self._table.item(row, _COL_NAME)
        name = name_item.text() if name_item else ""

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background:#ffffff; border:1px solid #e4eaf0;
                    border-radius:8px; padding:4px; font-size:13px; }
            QMenu::item { padding:6px 18px 6px 10px; border-radius:5px; }
            QMenu::item:selected { background:#dbeafe; color:#1d4ed8; }
            QMenu::separator { height:1px; background:#f0f4f8; margin:3px 6px; }
        """)

        chat_id = name_item.data(Qt.ItemDataRole.UserRole) if name_item else ""

        act_open      = menu.addAction("🌐  Открыть в браузере")
        act_open.setEnabled(link.startswith("http"))
        act_copy_link = menu.addAction("📋  Копировать ссылку")
        act_copy_link.setEnabled(bool(link))
        menu.addSeparator()
        act_copy_name = menu.addAction("📝  Копировать название")
        act_copy_name.setEnabled(bool(name))
        act_copy_id   = menu.addAction("🔢  Копировать chat_id")
        act_copy_id.setEnabled(bool(chat_id))

        chosen = menu.exec(self._table.viewport().mapToGlobal(pos))
        if chosen == act_open and link.startswith("http"):
            QDesktopServices.openUrl(QUrl(link))
        elif chosen == act_copy_link and link:
            QApplication.clipboard().setText(link)
        elif chosen == act_copy_name and name:
            QApplication.clipboard().setText(name)
        elif chosen == act_copy_id and chat_id:
            QApplication.clipboard().setText(chat_id)

    def _row_in_period(self, r: dict) -> bool:
        if self._period_days == 0:
            return True
        try:
            cutoff = datetime.now() - timedelta(days=self._period_days)
            return datetime.strptime(r["last_event"][:10], "%d.%m.%Y") >= cutoff
        except ValueError:
            return False

    # ── Экспорт в Excel ──────────────────────────────────────────

    def _export_excel(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Экспорт", "Библиотека openpyxl не установлена.")
            return

        # Имя файла включает период и дату
        period_file_labels = {0: "все", 1: "день", 7: "неделя", 30: "месяц"}
        period_file = period_file_labels.get(self._period_days, "все")
        date_str    = datetime.now().strftime("%Y%m%d_%H%M")
        default_name = f"статистика_групп_{period_file}_{date_str}.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить статистику", default_name, "Excel (*.xlsx)"
        )
        if not path:
            return

        period_label = self._PERIOD_LABELS.get(self._period_days)
        period_names = {1: "день", 7: "неделя", 30: "месяц"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Статистика групп"

        # ── Строка 1: заголовок отчёта ────────────────────────────
        period_str = period_names.get(self._period_days, "все периоды")
        generated  = datetime.now().strftime("%d.%m.%Y %H:%M")
        row_count  = self._table.rowCount()
        title_val  = f"Статистика групп · {period_str} · {generated} · строк: {row_count}"
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value     = title_val
        title_cell.font      = Font(bold=True, size=12)
        title_cell.fill      = PatternFill("solid", fgColor="4A6CF7")
        title_cell.font      = Font(bold=True, size=12, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # ── Строка 2: сводка периода ──────────────────────────────
        if period_label:
            members_active = sum(
                int(r["members"]) for r in self._all_rows
                if str(r["members"]).isdigit() and self._row_in_period(r)
            )
            summary_val = (
                f"{period_label.capitalize()}: активны {row_count} из {len(self._all_rows)} групп  |  "
                f"участников в активных: {members_active:,}".replace(",", "\u00a0")
            )
            ws.merge_cells("A2:D2")
            summ_cell = ws["A2"]
            summ_cell.value     = summary_val
            summ_cell.font      = Font(italic=True, size=10, color="374151")
            summ_cell.fill      = PatternFill("solid", fgColor="EFF2FF")
            summ_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 18
            data_start_row = 3
        else:
            data_start_row = 2

        # ── Заголовки таблицы ─────────────────────────────────────
        headers    = ["Название", "Участников", "Последняя активность",
                      "Δ Подписчики", "Ссылка"]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill("solid", fgColor="F1F5F9")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=col_idx, value=hdr)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        # ── Данные из таблицы (отфильтрованные строки) ────────────
        for row_idx in range(row_count):
            for col_idx in range(5):
                item  = self._table.item(row_idx, col_idx)
                value = item.text() if item else ""
                ws.cell(row=data_start_row + 1 + row_idx, column=col_idx + 1, value=value)

        for col in ws.columns:
            col_letter = None
            for c in col:
                if hasattr(c, "column_letter"):
                    col_letter = c.column_letter
                    break
            if not col_letter:
                continue
            max_len = max((len(str(c.value)) if hasattr(c, "value") and c.value else 0) for c in col)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        try:
            wb.save(path)
        except Exception as exc:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Экспорт", f"Ошибка сохранения: {exc}")
            return

        self._status_lbl.setText(f"✅  Экспортировано {row_count} строк → {Path(path).name}")

        # Всплывающее уведомление с кнопкой «Открыть»
        from PyQt6.QtWidgets import QMessageBox
        mb = QMessageBox(self)
        mb.setWindowTitle("Экспорт завершён")
        mb.setText(f"Файл сохранён:\n{Path(path).name}")
        mb.setInformativeText(f"Строк: {row_count}  ·  Период: {period_str}")
        mb.setIcon(QMessageBox.Icon.Information)
        open_btn = mb.addButton("📂  Открыть файл", QMessageBox.ButtonRole.ActionRole)
        mb.addButton("OK", QMessageBox.ButtonRole.AcceptRole)
        mb.exec()
        if mb.clickedButton() == open_btn:
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def _load_subscriber_history(self) -> None:
        """Загружает историю подписчиков: авто через бота или вручную из файла."""
        from dotenv import load_dotenv
        load_dotenv(get_env_path(), override=True)
        instance_id = os.getenv("MAX_ID_INSTANCE", "").strip()
        api_token   = os.getenv("MAX_API_TOKEN",   "").strip()

        if instance_id and api_token:
            self._load_history_via_bot(instance_id, api_token)
        else:
            self._load_history_from_file()

    def _load_history_via_bot(self, instance_id: str, token: str) -> None:
        """Запрашивает отчёт у бота через GREEN-API с выбором периода."""
        from PyQt6.QtWidgets import (
            QDialog, QVBoxLayout, QLabel, QDialogButtonBox,
            QButtonGroup, QRadioButton,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Запрос отчёта у бота")
        dlg.setMinimumWidth(280)
        lay = QVBoxLayout(dlg)
        lay.setSpacing(10)
        lay.addWidget(QLabel("Выберите период отчёта:"))

        bg     = QButtonGroup(dlg)
        radios: dict[QRadioButton, int] = {}
        for days, label in ((7, "7 дней"), (30, "30 дней"), (90, "90 дней")):
            rb = QRadioButton(label)
            if days == 30:
                rb.setChecked(True)
            bg.addButton(rb)
            lay.addWidget(rb)
            radios[rb] = days

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Запросить у бота")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Загрузить файл вручную")
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)

        result = dlg.exec()
        if result == QDialog.DialogCode.Rejected:
            # «Загрузить файл вручную»
            self._load_history_from_file()
            return

        days = next(d for rb, d in radios.items() if rb.isChecked())

        # Останавливаем старый воркер если был
        if self._bot_worker and self._bot_worker.isRunning():
            self._bot_worker.stop()
            self._bot_worker.quit()
            self._bot_worker.wait(2000)

        self._history_btn.setEnabled(False)
        self._history_btn.setText(f"📊 Загрузка {days}д…")

        self._bot_worker = _BotReportWorker(instance_id, token, days, parent=self)
        self._bot_worker.progress.connect(self._status_lbl.setText)
        self._bot_worker.done.connect(self._on_bot_report_done)
        self._bot_worker.failed.connect(self._on_bot_report_failed)
        self._bot_worker.finished.connect(
            lambda: self._history_btn.setEnabled(True)
        )
        self._bot_worker.start()

    def _on_bot_report_done(self, html: str) -> None:
        self._history_btn.setEnabled(True)
        self._parse_and_apply_history(html)

    def _on_bot_report_failed(self, msg: str) -> None:
        self._history_btn.setEnabled(True)
        self._history_btn.setText("📊 История")
        from PyQt6.QtWidgets import QMessageBox
        mb = QMessageBox(self)
        mb.setWindowTitle("Ошибка получения отчёта")
        mb.setText(msg)
        mb.setIcon(QMessageBox.Icon.Warning)
        manual_btn = mb.addButton("📂 Загрузить файл вручную", QMessageBox.ButtonRole.ActionRole)
        mb.addButton("Отмена", QMessageBox.ButtonRole.RejectRole)
        mb.exec()
        if mb.clickedButton() == manual_btn:
            self._load_history_from_file()

    def _load_history_from_file(self) -> None:
        """Открывает диалог выбора HTML-файла и парсит его."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Загрузить историю подписчиков", "",
            "HTML файлы (*.html *.htm);;Все файлы (*)"
        )
        if not path:
            return
        try:
            html = Path(path).read_text(encoding="utf-8", errors="replace")
            self._parse_and_apply_history(html)
        except Exception as exc:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Ошибка загрузки файла", str(exc))

    def _parse_and_apply_history(self, html: str) -> None:
        """Парсит HTML-отчёт подписчиков и обновляет таблицу с дельтой."""
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, "html.parser")

            period_text = ""
            for p in soup.find_all("p"):
                if "Период" in p.get_text():
                    period_text = p.get_text(strip=True)
                    break

            table = soup.find("table")
            if not table:
                raise ValueError("Таблица не найдена в HTML-файле")

            history: dict = {}
            skip_classes  = {"average-row", "total-row"}

            for tr in table.find_all("tr")[1:]:
                if set(tr.get("class", [])) & skip_classes:
                    continue
                tds = tr.find_all("td")
                if len(tds) < 4:
                    continue
                name = tds[1].get_text(strip=True)
                if not name:
                    continue

                values: list[int | None] = []
                for td in tds[3:]:
                    t = td.get_text(strip=True)
                    try:
                        values.append(int(t))
                    except ValueError:
                        values.append(None)

                non_null = [v for v in values if v is not None]
                if not non_null:
                    continue

                latest = non_null[0]
                oldest = non_null[-1]
                key = self._history_key(name)
                history[key] = {
                    "name":   name,
                    "latest": latest,
                    "oldest": oldest,
                    "delta":  latest - oldest,
                }

            self._history_data   = history
            self._history_period = period_text

            matched     = sum(1 for r in self._all_rows
                              if self._history_key(r["name"]) in history)
            grew        = sum(1 for v in history.values() if v["delta"] > 0)
            shrank      = sum(1 for v in history.values() if v["delta"] < 0)
            total_delta = sum(v["delta"] for v in history.values())
            sign        = "+" if total_delta >= 0 else ""

            self._history_banner.setText(
                f"📊  {period_text}  ·  "
                f"групп: {len(history)}  ·  совпало: {matched}  ·  "
                f"растут: {grew} ↑  падают: {shrank} ↓  "
                f"итого: {sign}{total_delta:,}".replace(",", "\u00a0")
            )
            self._history_banner.show()
            self._history_btn.setText("📊 История ✓")
            self._apply_filter()
            self._status_lbl.setText(
                f"История загружена: {len(history)} групп · {matched} совпало"
            )

        except Exception as exc:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Ошибка разбора отчёта", str(exc))

    # ── Тема / закрытие ──────────────────────────────────────────

    def _shutdown(self) -> None:
        self._auto_timer.stop()
        if self._worker and self._worker.isRunning():
            self._worker._stop = True
            self._worker.quit()
            self._worker.wait(15000)
        if self._bot_worker and self._bot_worker.isRunning():
            self._bot_worker.stop()
            self._bot_worker.quit()
            self._bot_worker.wait(3000)

    def closeEvent(self, event) -> None:
        self._shutdown()
        super().closeEvent(event)

    def set_dark(self, dark: bool) -> None:
        self._apply_styles(dark=dark)

    def _apply_styles(self, dark: bool = False) -> None:
        if dark:
            self.setStyleSheet("""
                StatsPanel { background: #1e1e2e; }
                QLabel#statsPanelTitle { font-size: 18px; font-weight: 700; color: #d0d0f0; }
                QLabel#statsLastRefresh { font-size: 11px; color: #6868aa; padding-right: 8px; }
                QPushButton#statsToggleBtn {
                    min-height:0; font-size:13px; font-weight:600; padding:4px 20px;
                    border:1px solid #3a3a55; background:#2d2d45; color:#8888aa;
                }
                QPushButton#statsToggleBtn:first-child { border-radius: 8px 0 0 8px; }
                QPushButton#statsToggleBtn:last-child  { border-radius: 0 8px 8px 0; }
                QPushButton#statsToggleBtnActive {
                    min-height:0; font-size:13px; font-weight:600; padding:4px 20px;
                    border:1px solid #4a6cf7; background:#4a6cf7; color:#ffffff;
                }
                QPushButton#statsToggleBtnActive:first-child { border-radius: 8px 0 0 8px; }
                QPushButton#statsToggleBtnActive:last-child  { border-radius: 0 8px 8px 0; }
                QPushButton#statsPeriodBtn {
                    min-height:0; font-size:12px; padding:2px 14px;
                    border:1px solid #3a3a55; border-radius:6px; background:#2d2d45; color:#8888aa;
                }
                QPushButton#statsPeriodBtnActive {
                    min-height:0; font-size:12px; padding:2px 14px;
                    border:1px solid #4a6cf7; border-radius:6px; background:#2a3a6a; color:#8899ff;
                }
                QPushButton#statsRefreshBtn {
                    min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 16px;
                    border-radius: 7px; border: 1px solid #3a3a55; background: #2d2d45; color: #c8c8e0;
                }
                QPushButton#statsRefreshBtn:hover { background: #1e2a5a; border-color: #4a6cf7; color: #8899ff; }
                QPushButton#statsRefreshBtn:disabled { color: #5a5a88; }
                QFrame#statsSummary { background: #252535; border: 1px solid #3a3a55; border-radius: 10px; }
                QLabel#statsStatValue { font-size: 24px; font-weight: 700; color: #a0c0ff; }
                QLabel#statsStatLabel { font-size: 11px; color: #6868aa; font-weight: 500; }
                QLineEdit#statsSearch {
                    font-size: 13px; padding: 7px 12px; border: 1px solid #3a3a55;
                    border-radius: 8px; background: #2a2a3e; color: #d0d0f0;
                }
                QTableWidget#statsTable {
                    border: 1px solid #3a3a55; border-radius: 8px; background: #252535;
                    alternate-background-color: #222232; gridline-color: #2d2d45;
                    font-size: 12px; color: #d0d0f0;
                }
                QTableWidget#statsTable QHeaderView::section {
                    background: #222232; color: #7878aa; font-size: 11px; font-weight: 700;
                    padding: 6px 10px; border: none; border-bottom: 2px solid #3a3a55;
                    text-transform: uppercase; letter-spacing: 0.4px;
                }
                QTableWidget#statsTable::item:selected { background: #1e3a8a; color: #e0e0ff; }
                QLabel#statsStatus { font-size: 11px; color: #6868aa; padding: 2px 0; }
                QPushButton#statsExportBtn {
                    min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 14px;
                    border-radius: 7px; border: 1px solid #2a5a3e; background: #1a3a2e; color: #4ade80;
                }
                QPushButton#statsExportBtn:hover { background: #1e4a36; border-color: #22c55e; }
                QPushButton#statsExportBtn:disabled { color: #5a5a88; background: #222232; border-color: #333344; }
                QPushButton#statsDeadBtn {
                    min-height: 0; font-size: 12px; padding: 5px 12px; border-radius: 7px;
                    border: 1px solid #6b2020; background: #2a1a1a; color: #f87171; font-weight: 600;
                }
                QPushButton#statsDeadBtn:checked { background: #dc2626; color: #ffffff; border-color: #dc2626; }
                QPushButton#statsDeadBtn:disabled { color: #5a5a88; background: #222232; border-color: #333344; }
                QLabel#statsCacheBanner {
                    font-size: 12px; font-weight: 600; color: #fcd34d;
                    background: #3a2a00; border: 1px solid #78580a;
                    border-radius: 7px; padding: 6px 12px;
                }
                QLabel#statsHistoryBanner {
                    font-size: 12px; font-weight: 600; color: #4ade80;
                    background: #0a2a1a; border: 1px solid #166534;
                    border-radius: 7px; padding: 6px 12px;
                }
                QPushButton#statsHistoryBtn {
                    min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 14px;
                    border-radius: 7px; border: 1px solid #1a3a5a; background: #0f2a40; color: #60a5fa;
                }
                QPushButton#statsHistoryBtn:hover { background: #1a3a5a; border-color: #3b82f6; }
                QProgressBar#statsProgressBar {
                    border: none; border-radius: 3px; background: #2d2d45;
                }
                QProgressBar#statsProgressBar::chunk {
                    border-radius: 3px; background: qlineargradient(
                        x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4a6cf7, stop:1 #7c3aed);
                }
                QLabel#statsPeriodSummary {
                    font-size: 12px; color: #8899ff;
                    background: #1e2540; border: 1px solid #2a3a6a;
                    border-radius: 6px; padding: 5px 12px;
                }
            """)
            return
        self.setStyleSheet("""
            StatsPanel { background: #f3f4f6; }
            QLabel#statsPanelTitle { font-size: 18px; font-weight: 700; color: #1a1a2e; }
            QLabel#statsLastRefresh { font-size: 11px; color: #9ca3af; padding-right: 8px; }
            QPushButton#statsToggleBtn {
                min-height:0; font-size:13px; font-weight:600; padding:4px 20px;
                border:1px solid #c7d0db; background:#eef2f7; color:#6b7280;
            }
            QPushButton#statsToggleBtn:first-child { border-radius: 8px 0 0 8px; }
            QPushButton#statsToggleBtn:last-child  { border-radius: 0 8px 8px 0; }
            QPushButton#statsToggleBtnActive {
                min-height:0; font-size:13px; font-weight:600; padding:4px 20px;
                border:1px solid #4a6cf7; background:#4a6cf7; color:#ffffff;
            }
            QPushButton#statsToggleBtnActive:first-child { border-radius: 8px 0 0 8px; }
            QPushButton#statsToggleBtnActive:last-child  { border-radius: 0 8px 8px 0; }
            QPushButton#statsPeriodBtn {
                min-height:0; font-size:12px; padding:2px 14px;
                border:1px solid #c7d0db; border-radius:6px; background:#eef2f7; color:#6b7280;
            }
            QPushButton#statsPeriodBtnActive {
                min-height:0; font-size:12px; padding:2px 14px;
                border:1px solid #4a6cf7; border-radius:6px; background:#eff2ff; color:#4a6cf7;
            }
            QPushButton#statsRefreshBtn {
                min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 16px;
                border-radius: 7px; border: 1px solid #c7d0db; background: #eef2f7; color: #334155;
            }
            QPushButton#statsRefreshBtn:hover { background: #dbeafe; border-color: #2d6cdf; color: #1d4ed8; }
            QPushButton#statsRefreshBtn:disabled { color: #9ca3af; }
            QFrame#statsSummary { background: #ffffff; border: 1px solid #e4eaf0; border-radius: 10px; }
            QLabel#statsStatValue { font-size: 24px; font-weight: 700; color: #1e3a5f; }
            QLabel#statsStatLabel { font-size: 11px; color: #9ca3af; font-weight: 500; }
            QLineEdit#statsSearch {
                font-size: 13px; padding: 7px 12px; border: 1px solid #c7d0db;
                border-radius: 8px; background: #ffffff; color: #1a1a2e;
            }
            QTableWidget#statsTable {
                border: 1px solid #e4eaf0; border-radius: 8px; background: #ffffff;
                alternate-background-color: #f8fafc; gridline-color: #f0f4f8;
                font-size: 12px; color: #1a1a2e;
            }
            QTableWidget#statsTable QHeaderView::section {
                background: #f1f5f9; color: #64748b; font-size: 11px; font-weight: 700;
                padding: 6px 10px; border: none; border-bottom: 2px solid #e2e8f0;
                text-transform: uppercase; letter-spacing: 0.4px;
            }
            QTableWidget#statsTable::item:selected { background: #dbeafe; color: #1e3a5f; }
            QLabel#statsStatus { font-size: 11px; color: #9ca3af; padding: 2px 0; }
            QPushButton#statsExportBtn {
                min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 14px;
                border-radius: 7px; border: 1px solid #a7f3d0; background: #ecfdf5; color: #065f46;
            }
            QPushButton#statsExportBtn:hover { background: #d1fae5; border-color: #34d399; }
            QPushButton#statsExportBtn:disabled { color: #9ca3af; background: #f3f4f6; border-color: #e5e7eb; }
            QPushButton#statsDeadBtn {
                min-height: 0; font-size: 12px; padding: 5px 12px; border-radius: 7px;
                border: 1px solid #f5c6c6; background: #fff5f5; color: #dc2626; font-weight: 600;
            }
            QPushButton#statsDeadBtn:checked { background: #dc2626; color: #ffffff; border-color: #dc2626; }
            QPushButton#statsDeadBtn:disabled { color: #ccc; background: #f9f9f9; border-color: #e5e7eb; }
            QLabel#statsCacheBanner {
                font-size: 12px; font-weight: 600; color: #92400e;
                background: #fef3c7; border: 1px solid #fcd34d;
                border-radius: 7px; padding: 6px 12px;
            }
            QLabel#statsHistoryBanner {
                font-size: 12px; font-weight: 600; color: #15803d;
                background: #f0fdf4; border: 1px solid #86efac;
                border-radius: 7px; padding: 6px 12px;
            }
            QPushButton#statsHistoryBtn {
                min-height: 0; font-size: 13px; font-weight: 600; padding: 4px 14px;
                border-radius: 7px; border: 1px solid #bfdbfe; background: #eff6ff; color: #1d4ed8;
            }
            QPushButton#statsHistoryBtn:hover { background: #dbeafe; border-color: #3b82f6; }
            QProgressBar#statsProgressBar {
                border: none; border-radius: 3px; background: #e2e8f0;
            }
            QProgressBar#statsProgressBar::chunk {
                border-radius: 3px; background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #8b5cf6);
            }
            QLabel#statsPeriodSummary {
                font-size: 12px; color: #374151;
                background: #eff2ff; border: 1px solid #c7d5fb;
                border-radius: 6px; padding: 5px 12px;
            }
        """)


class _NumItem(QTableWidgetItem):
    """Элемент таблицы с числовой сортировкой."""

    def __lt__(self, other: QTableWidgetItem) -> bool:
        try:
            return int(self.text()) < int(other.text())
        except ValueError:
            return super().__lt__(other)
